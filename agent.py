"""
Opportunities Newsletter agent.

Runs on a monthly schedule. Uses a Claude agent loop to:
  1. Search Sam's Gmail for the most recent "Opportunities Newsletter" thread
  2. Verify which programs from the skill's organisation list are currently open
  3. Return structured JSON describing the next newsletter
  4. Build email body + xlsx spreadsheet
  5. Save the email as a Gmail draft (with the spreadsheet attached)
  6. Email Sam a notification that the draft is ready for review

Sam reviews the draft in Gmail, edits as needed, and sends it.
"""
from __future__ import annotations

import base64
import json
import os
import sys
import time
import urllib.request
from datetime import date
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
from html import escape
from pathlib import Path

from anthropic import Anthropic
from google.auth.transport.requests import Request as GoogleRequest
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# --- Configuration --------------------------------------------------------

SAM_EMAIL = os.environ.get("SAM_EMAIL", "samfoxanu@gmail.com")
MODEL = os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-6")
CLAUDE_CALL_DELAY_SECONDS = int(os.environ.get("CLAUDE_CALL_DELAY_SECONDS", "65"))
SKILL_DIR = Path(__file__).parent / "skill"
OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)
TODAY = date.today().isoformat()

GMAIL_SCOPES = [
    "https://www.googleapis.com/auth/gmail.compose",
    "https://www.googleapis.com/auth/gmail.send",
    "https://www.googleapis.com/auth/gmail.readonly",
]


# --- Gmail helpers --------------------------------------------------------

def gmail_service():
    """Build a Gmail service from a stored refresh token."""
    creds = Credentials(
        token=None,
        refresh_token=os.environ["GMAIL_REFRESH_TOKEN"],
        client_id=os.environ["GMAIL_CLIENT_ID"],
        client_secret=os.environ["GMAIL_CLIENT_SECRET"],
        token_uri="https://oauth2.googleapis.com/token",
        scopes=GMAIL_SCOPES,
    )
    creds.refresh(GoogleRequest())
    return build("gmail", "v1", credentials=creds, cache_discovery=False)


def gmail_search(query: str, max_results: int = 5) -> list[dict]:
    svc = gmail_service()
    res = svc.users().threads().list(userId="me", q=query, maxResults=max_results).execute()
    return res.get("threads", [])


def gmail_get_thread(thread_id: str) -> str:
    svc = gmail_service()
    thread = svc.users().threads().get(userId="me", id=thread_id, format="full").execute()
    snippets = []
    for msg in thread.get("messages", []):
        snippets.append(extract_body(msg["payload"]))
    return "\n\n---\n\n".join(snippets)


def extract_body(payload) -> str:
    if payload.get("body", {}).get("data"):
        return base64.urlsafe_b64decode(payload["body"]["data"]).decode("utf-8", "replace")
    parts = payload.get("parts", []) or []
    out = []
    for p in parts:
        if p.get("mimeType") == "text/plain":
            out.append(extract_body(p))
        elif p.get("parts"):
            out.append(extract_body(p))
    return "\n".join(out)


def create_gmail_draft(
    subject: str,
    body: str,
    attachment_path: Path | None = None,
    html_body: str | None = None,
) -> str:
    msg = MIMEMultipart("mixed")
    msg["to"] = SAM_EMAIL
    msg["from"] = SAM_EMAIL
    msg["subject"] = subject

    if html_body:
        alternative = MIMEMultipart("alternative")
        alternative.attach(MIMEText(body, "plain", "utf-8"))
        alternative.attach(MIMEText(html_body, "html", "utf-8"))
        msg.attach(alternative)
    else:
        msg.attach(MIMEText(body, "plain", "utf-8"))

    if attachment_path and attachment_path.exists():
        part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(attachment_path.read_bytes())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{attachment_path.name}"')
        msg.attach(part)
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
    svc = gmail_service()
    draft = svc.users().drafts().create(userId="me", body={"message": {"raw": raw}}).execute()
    return draft["id"]


def send_notification_email(draft_id: str, n_programs: int):
    body = (
        f"Your monthly Opportunities Newsletter draft is ready for review.\n\n"
        f"Programs included: {n_programs}\n"
        f"Date generated: {TODAY}\n\n"
        f"Reminder: update Google Contacts label for all new signups.\n\n"
        f"Open Gmail Drafts to review, edit, and send when ready.\n"
        f"https://mail.google.com/mail/u/0/#drafts/{draft_id}\n"
    )
    msg = MIMEText(body, "plain", "utf-8")
    msg["to"] = SAM_EMAIL
    msg["from"] = SAM_EMAIL
    msg["subject"] = f"[Bot] Opportunities Newsletter draft ready — {TODAY}"
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
    svc = gmail_service()
    svc.users().messages().send(userId="me", body={"raw": raw}).execute()


# --- Web fetch ------------------------------------------------------------

def web_fetch(url: str, max_chars: int = 2500) -> str:
    try:
        req = urllib.request.Request(
            url,
            headers={"User-Agent": "Mozilla/5.0 (compatible; OpportunitiesNewsletterBot/1.0)"},
        )
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = resp.read(200_000).decode("utf-8", "replace")
        # Crude tag stripping — Claude can parse what's left
        import re
        text = re.sub(r"<script[^>]*>.*?</script>", "", data, flags=re.S | re.I)
        text = re.sub(r"<style[^>]*>.*?</style>", "", text, flags=re.S | re.I)
        text = re.sub(r"<[^>]+>", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
        return text[:max_chars]
    except Exception as exc:
        return f"[FETCH ERROR: {exc}]"


# --- Spreadsheet ----------------------------------------------------------

def build_spreadsheet(programs: list[dict], out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Open Programs"
    headers = [
        "Country", "Organisation", "Program name", "Type",
        "Deadline", "Duration", "Paid?",
        "International eligibility", "Notes", "Apply URL",
    ]
    ws.append(headers)
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="305496")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    data_font = Font(name="Calibri", size=11)
    for p in programs:
        ws.append([
            p.get("country", ""),
            p.get("organisation", ""),
            p.get("program_name", ""),
            p.get("type", ""),
            p.get("deadline", ""),
            p.get("duration", ""),
            p.get("paid", ""),
            p.get("international", ""),
            p.get("notes", ""),
            p.get("url", ""),
        ])
    for r in range(2, len(programs) + 2):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border
    widths = {"A": 16, "B": 28, "C": 38, "D": 16, "E": 22, "F": 26, "G": 28, "H": 22, "I": 50, "J": 56}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 30
    for r in range(2, len(programs) + 2):
        ws.row_dimensions[r].height = 50
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(out_path)


# --- Email body assembly --------------------------------------------------

COUNTRY_FLAGS = {
    "Australia": "🇦🇺", "New Zealand": "🇳🇿", "United Kingdom": "🇬🇧",
    "Europe": "🇪🇺", "United States": "🇺🇸", "Canada": "🇨🇦",
    "India": "🌏", "Asia": "🌏", "Latin America": "🌎", "Mexico": "🌎",
    "Middle East": "🌍", "Africa": "🌍", "Global / Online": "🌐",
}
COUNTRY_ORDER = [
    "Australia", "New Zealand", "United Kingdom", "Europe",
    "United States", "Canada", "Asia", "India", "Latin America", "Mexico",
    "Middle East", "Africa", "Global / Online",
]


def build_email_body(payload: dict) -> str:
    intro = payload.get("intro_block", "").strip()
    closing = payload.get("closing_block", "").strip()
    programs = payload.get("programs", [])

    # Group by country, ordered
    by_country: dict[str, list[dict]] = {}
    for p in programs:
        by_country.setdefault(p["country"], []).append(p)

    sections = []
    seen = set()
    for country in COUNTRY_ORDER:
        if country in by_country:
            sections.append(format_section(country, by_country[country]))
            seen.add(country)
    # Any country we didn't anticipate gets appended at the end
    for country, items in by_country.items():
        if country not in seen:
            sections.append(format_section(country, items))

    body = intro + "\n\n---\n\n" + "\n\n---\n\n".join(sections) + "\n\n---\n\n" + closing
    return body


def format_section(country: str, items: list[dict]) -> str:
    flag = COUNTRY_FLAGS.get(country, "🌐")
    header = f"{flag} {country.upper()}"
    blocks = [header, ""]
    for p in items:
        header_line = f"{p['organisation']} — {p['program_name']} | {p['deadline']}"
        blocks.append(header_line)
        blocks.append("")
        blocks.append(p["description_paragraph"].strip())
        blocks.append("")
        blocks.append(p["url"])
        blocks.append("")
    return "\n".join(blocks).rstrip()


def body_to_html(body: str) -> str:
    parts = [
        '<div style="font-family: Arial, Helvetica, sans-serif; font-size: 14px; line-height: 1.5; color: #202124;">'
    ]
    for raw_line in body.splitlines():
        line = raw_line.strip()
        if not line:
            parts.append('<div style="height: 10px;"></div>')
            continue
        if line == "---":
            parts.append(
                '<div style="height: 18px;"></div>'
                '<hr style="border: 0; border-top: 1px solid #dadce0; margin: 0 0 22px;">'
            )
            continue
        if is_section_header(line):
            parts.append(
                f'<h2 style="font-size: 24px; line-height: 1.3; font-weight: 700; '
                f'margin: 24px 0 10px; color: #111827;">{escape(line)}</h2>'
            )
            continue
        if " — " in line and " | " in line:
            parts.append(
                f'<p style="font-size: 17px; font-weight: 700; margin: 14px 0 4px;">'
                f'{escape(line)}</p>'
            )
            continue
        if line.startswith("http://") or line.startswith("https://"):
            safe_url = escape(line, quote=True)
            parts.append(f'<p style="margin: 4px 0 12px;"><a href="{safe_url}">{escape(line)}</a></p>')
            continue
        parts.append(f'<p style="margin: 0 0 8px;">{escape(line)}</p>')
    parts.append("</div>")
    return "\n".join(parts)


def is_section_header(line: str) -> bool:
    text = line
    for flag in COUNTRY_FLAGS.values():
        if text.startswith(flag):
            text = text[len(flag):].strip()
            break
    return bool(text) and text == text.upper() and any(ch.isalpha() for ch in text)


# --- Agent loop -----------------------------------------------------------

def load_skill() -> str:
    parts = []
    for path in sorted(SKILL_DIR.rglob("*.md")):
        rel = path.relative_to(SKILL_DIR)
        if rel.as_posix() == "references/organisations.md":
            parts.append(
                "=== references/organisations.md ===\n\n"
                "The full organisation/program source list is available through "
                "the read_reference_file tool. Read it before checking live pages."
            )
            continue
        parts.append(f"=== {rel} ===\n\n{path.read_text(encoding='utf-8')}")
    return "\n\n".join(parts)


def read_reference_file(filename: str) -> str:
    safe_name = Path(filename).name
    path = SKILL_DIR / "references" / safe_name
    allowed = {
        "organisations.md",
        "known-traps.md",
        "spreadsheet-schema.md",
        "template.md",
    }
    if safe_name not in allowed or not path.exists():
        return f"[UNKNOWN REFERENCE FILE: {filename}]"
    return path.read_text(encoding="utf-8")[:24000]


SYSTEM_PROMPT = """You are an assistant that drafts Sam Fox's monthly "Opportunities Newsletter".

Below are the skill files that define exactly how to do this. Follow them carefully.

{skill_content}

Today's date is {today}.

YOUR JOB:

1. Call gmail_search with query `subject:"Opportunities Newsletter"` to find what's been featured recently.
2. Optionally call gmail_get_thread on the most recent thread to read what was in the last newsletter.
3. Call read_reference_file for `organisations.md`, then for every Tier 1 program listed there, call web_fetch on its specific application URL and decide whether it is CURRENTLY OPEN today.
4. For any Tier 2 program you think is likely to be open this cycle, also web_fetch it.
5. When you have a verified list of currently-open programs, call submit_newsletter with the final structured payload.

CRITICAL RULES:
- Closed programs are EXCLUDED — not in the email body, not in the spreadsheet.
- When uncertain whether a program is open, OMIT IT. Better a shorter accurate newsletter than a long one with broken listings.
- Never invent a deadline. Use the hedging language from template.md: "Rolling", "Apply early", "Contact for current application window", "Applications open".
- Do not use the word "genuinely".
- Match the country order, type order, header format, and tone from template.md.

Once you call submit_newsletter you are done.
"""

TOOLS = [
    {
        "name": "read_reference_file",
        "description": "Read one newsletter reference file, especially organisations.md.",
        "input_schema": {
            "type": "object",
            "properties": {"filename": {"type": "string"}},
            "required": ["filename"],
        },
    },
    {
        "name": "gmail_search",
        "description": "Search Sam's Gmail. Returns a list of threads matching the query.",
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "Gmail query string."},
                "max_results": {"type": "integer", "default": 5},
            },
            "required": ["query"],
        },
    },
    {
        "name": "gmail_get_thread",
        "description": "Read the full plaintext body of a Gmail thread.",
        "input_schema": {
            "type": "object",
            "properties": {"thread_id": {"type": "string"}},
            "required": ["thread_id"],
        },
    },
    {
        "name": "web_fetch",
        "description": "Fetch the visible text of a web page. Use for verifying live application pages.",
        "input_schema": {
            "type": "object",
            "properties": {"url": {"type": "string"}},
            "required": ["url"],
        },
    },
    {
        "name": "submit_newsletter",
        "description": (
            "Submit the final newsletter payload. Call this exactly once when you are done. "
            "The payload becomes the Gmail draft and the spreadsheet attachment."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "intro_block": {
                    "type": "string",
                    "description": (
                        "Full intro text including greeting, signup line, framing paragraph, "
                        "and the spreadsheet-attached note. Follows template.md exactly."
                    ),
                },
                "programs": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "country": {"type": "string"},
                            "organisation": {"type": "string"},
                            "program_name": {"type": "string"},
                            "type": {
                                "type": "string",
                                "enum": [
                                    "Internship", "Conference", "Fellowship",
                                    "Essay competition", "Online course",
                                    "Seminar", "Scholarship", "Student program", "Other",
                                ],
                            },
                            "deadline": {"type": "string"},
                            "duration": {"type": "string"},
                            "paid": {"type": "string"},
                            "international": {
                                "type": "string",
                                "enum": ["Yes", "No", "Some restrictions"],
                            },
                            "notes": {"type": "string"},
                            "url": {"type": "string"},
                            "description_paragraph": {
                                "type": "string",
                                "description": "The 3-5 sentence email-body description paragraph.",
                            },
                        },
                        "required": [
                            "country", "organisation", "program_name", "type",
                            "deadline", "duration", "paid", "international",
                            "url", "description_paragraph",
                        ],
                    },
                },
                "closing_block": {
                    "type": "string",
                    "description": "Closing line(s) plus the 'Best, Sam' sign-off.",
                },
            },
            "required": ["intro_block", "programs", "closing_block"],
        },
    },
]


def run_tool(name: str, args: dict) -> str:
    try:
        if name == "read_reference_file":
            return read_reference_file(args["filename"])
        if name == "gmail_search":
            results = gmail_search(args["query"], args.get("max_results", 5))
            return json.dumps([{"id": t["id"], "snippet": t.get("snippet", "")} for t in results])[:4000]
        if name == "gmail_get_thread":
            return gmail_get_thread(args["thread_id"])[:5000]
        if name == "web_fetch":
            return web_fetch(args["url"])
        return f"[Unknown tool: {name}]"
    except Exception as exc:
        return f"[TOOL ERROR: {exc}]"


def run_agent() -> dict:
    client = Anthropic()
    system = SYSTEM_PROMPT.format(skill_content=load_skill(), today=TODAY)
    messages = [{"role": "user", "content": "Please draft this month's Opportunities Newsletter."}]

    for step in range(60):  # generous cap
        if step:
            print(f"[{TODAY}] Waiting {CLAUDE_CALL_DELAY_SECONDS}s to stay under API rate limits...")
            time.sleep(CLAUDE_CALL_DELAY_SECONDS)
        resp = client.messages.create(
            model=MODEL,
            max_tokens=8000,
            system=system,
            tools=TOOLS,
            messages=messages,
        )
        messages.append({"role": "assistant", "content": resp.content})

        if resp.stop_reason != "tool_use":
            raise RuntimeError(f"Agent stopped without submitting newsletter. stop_reason={resp.stop_reason}")

        tool_results = []
        submitted = None
        for block in resp.content:
            if block.type == "tool_use":
                if block.name == "submit_newsletter":
                    submitted = block.input
                    tool_results.append({
                        "type": "tool_result",
                        "tool_use_id": block.id,
                        "content": "Newsletter submitted. Thank you.",
                    })
                else:
                    result = run_tool(block.name, block.input)
                    tool_results.append({
                        "type": "tool_result",
                        "tool_use_id": block.id,
                        "content": result,
                    })

        messages.append({"role": "user", "content": tool_results})
        if submitted:
            return submitted

    raise RuntimeError("Agent did not submit a newsletter within step budget.")


# --- Main -----------------------------------------------------------------

def main():
    print(f"[{TODAY}] Starting newsletter draft generation...")
    payload = run_agent()
    n = len(payload.get("programs", []))
    print(f"[{TODAY}] Agent returned {n} programs.")

    spreadsheet_path = OUTPUT_DIR / f"opportunities-newsletter-{TODAY}.xlsx"
    build_spreadsheet(payload["programs"], spreadsheet_path)
    print(f"[{TODAY}] Spreadsheet saved: {spreadsheet_path}")

    body = build_email_body(payload)
    draft_id = create_gmail_draft("Opportunities Newsletter", body, spreadsheet_path, body_to_html(body))
    print(f"[{TODAY}] Gmail draft created: {draft_id}")

    send_notification_email(draft_id, n)
    print(f"[{TODAY}] Notification email sent to {SAM_EMAIL}.")


if __name__ == "__main__":
    main()
